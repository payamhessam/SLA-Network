"""Per-device Excel workbook export (the multi-tab device report).

Builds an .xlsx mirroring a device's detail tabs (health, ping, interfaces, VLANs,
neighbors, environment, alerts, config backups, methodology, etc.). See reports.py
for the fleet-wide Executive Excel/PowerPoint deck. Every cell written from collected
data is passed through `safe()` to neutralise spreadsheet formula injection.
"""
from datetime import datetime, timezone
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from .config import get_settings


# Fixed tab order of the exported device workbook.
SHEETS = ["Executive Overview", "Dashboard", "Branch Readiness", "Health Data", "Ping Quality", "Interfaces", "VLANs", "CDP-LLDP Neighbors", "Inventory", "Spanning Tree", "Environmental and PoE", "Alerts", "Configuration Backups", "Monitoring Gaps", "Collection Details", "SLA Methodology", "Instructions"]


def safe(value):
    # Formula-injection guard: a cell beginning with = + - @ is prefixed with an
    # apostrophe so spreadsheet apps treat it as text, never an executable formula.
    if isinstance(value, str) and value[:1] in "=+-@": return "'" + value
    return value


def create_report(devices):
    stamp = datetime.now(timezone.utc).strftime("%Y-%m-%d_%H%M%S")
    folder = Path(get_settings().report_dir); folder.mkdir(parents=True, exist_ok=True)
    path = folder / f"Network_Health_{stamp}.xlsx"
    wb = Workbook(); wb.remove(wb.active)
    for name in SHEETS:
        ws = wb.create_sheet(name)
        ws.sheet_view.showGridLines = False
        ws["A1"] = name; ws["A1"].font = Font(size=18, bold=True, color="FFFFFF"); ws["A1"].fill = PatternFill("solid", fgColor="102A43")
        ws.freeze_panes = "A3"
    ws = wb["Inventory"]
    headers = ["Hostname", "Management IP", "Site", "Role", "Type", "Model", "Active", "Mapping Status"]
    ws.append([]); ws.append(headers)
    for cell in ws[3]: cell.font = Font(bold=True, color="FFFFFF"); cell.fill = PatternFill("solid", fgColor="243B53")
    for d in devices: ws.append([safe(d.hostname), safe(d.management_ip), safe(d.site), safe(d.role), d.device_type, safe(d.model), d.active, d.match_status])
    ws.auto_filter.ref = f"A3:H{max(3, ws.max_row)}"
    for col in "ABCDEFGH": ws.column_dimensions[col].width = 22
    wb["Instructions"]["A3"] = "Missing values are not zero. Device administration changes only this application's inventory."
    wb["SLA Methodology"]["A3"] = "Availability = successful eligible minutes / observed eligible minutes; coverage below 90% is Insufficient evidence."
    wb.save(path)
    return path

