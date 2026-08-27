"""Access Point inventory and derived online/offline status.

Manages the Cisco 9120/9130 access-point inventory (add/import/edit/remove) and,
because LogicMonitor does not monitor these APs directly, derives each AP's status
every couple of hours: an AP is considered Online when its name or MAC appears in the
CDP/LLDP neighbor table of any monitored switch, Offline when it does not — recording
the connected switch, interface, and last-seen time. Status is inferred evidence,
never a fabricated up/down flag.
"""
import asyncio
import hashlib
import io
import logging
import re
import time
import uuid
import zipfile
from collections import Counter
from datetime import datetime, timezone
from ipaddress import ip_address

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from openpyxl import load_workbook
from sqlalchemy import delete, func, or_, select
from sqlalchemy.orm import Session

from .auth import administrator, current_user
from .config import get_settings
from .db import AccessPointImport, AccessPointInventory, AuditEvent, Device, SessionLocal, Site, Snapshot, session

logger = logging.getLogger(__name__)

router=APIRouter(prefix="/api/v1/access-points",tags=["Access Points"])
REQUIRED=["AP Name","AP Model","IP Address","AP Radio MAC","Ethernet MAC","Serial Number","Site Tag"]
MAC_RE=re.compile(r"^[0-9a-f]{12}$")
MAX_AP_UPLOAD_BYTES = 10 * 1024 * 1024
MAX_AP_IMPORT_ROWS = 5_000
MAX_XLSX_UNCOMPRESSED_BYTES = 100 * 1024 * 1024
MAX_XLSX_COMPRESSION_RATIO = 100


def norm_mac(value): return re.sub(r"[^0-9a-f]","",str(value or "").lower())
def display_mac(value):
    raw=norm_mac(value)
    return ":".join(raw[i:i+2] for i in range(0,12,2)) if len(raw)==12 else str(value or "").strip()


def _validate_workbook_bytes(content: bytes) -> None:
    """Reject oversized and highly-compressed workbook payloads before openpyxl expands them."""
    if len(content) > MAX_AP_UPLOAD_BYTES:
        raise HTTPException(413, "The access-point workbook must be 10 MB or smaller")
    try:
        with zipfile.ZipFile(io.BytesIO(content)) as archive:
            entries = archive.infolist()
            expanded = sum(entry.file_size for entry in entries)
            suspicious = any(entry.compress_size and entry.file_size / entry.compress_size > MAX_XLSX_COMPRESSION_RATIO for entry in entries)
    except zipfile.BadZipFile:
        raise HTTPException(422, "The uploaded workbook is not a valid Excel file")
    if expanded > MAX_XLSX_UNCOMPRESSED_BYTES or suspicious:
        raise HTTPException(413, "The access-point workbook expands beyond the safe processing limit")


def latest_neighbors(db):
    """Read only the latest snapshot per switch, avoiding a full-history scan per request."""
    latest = select(Snapshot.device_id.label("device_id"), func.max(Snapshot.collected_at).label("collected_at")).group_by(Snapshot.device_id).subquery()
    stmt = select(Snapshot, Device.hostname).join(latest, (Snapshot.device_id == latest.c.device_id) & (Snapshot.collected_at == latest.c.collected_at)).outerjoin(Device, Device.id == Snapshot.device_id)
    found=[];freshest=None
    for snap, hostname in db.execute(stmt):
        if freshest is None or snap.collected_at > freshest:
            freshest = snap.collected_at
        for row in (snap.details or {}).get("tables",{}).get("CDP-LLDP Neighbors",[]):
            found.append({"name":str(row.get("Neighbor") or "").strip().lower().split(".")[0],"mac":norm_mac(row.get("Ethernet MAC") or row.get("Management MAC") or row.get("Chassis ID")),"switch":hostname,"interface":row.get("Local Interface"),"detected":snap.collected_at})
    return found, freshest


def neighbor_evidence_is_fresh(collected_at) -> bool:
    if not collected_at:
        return False
    if collected_at.tzinfo is None:
        collected_at = collected_at.replace(tzinfo=timezone.utc)
    limit = max(60, get_settings().switch_collection_interval_minutes * 2)
    return (datetime.now(timezone.utc) - collected_at).total_seconds() <= limit * 60


def detected(ap,neighbors):
    for candidate in neighbors:
        if (ap.ethernet_mac and candidate["mac"]==norm_mac(ap.ethernet_mac)) or candidate["name"]==ap.ap_name.lower().split(".")[0] or (ap.radio_mac and candidate["mac"]==norm_mac(ap.radio_mac)):
            return candidate
    return None


def row_json(ap,neighbors,evidence_fresh):
    match=detected(ap,neighbors)
    status = "Online" if match else ("Offline" if evidence_fresh else "Unknown")
    return {"id":ap.id,"status":status,"ap_name":ap.ap_name,"site_code":ap.site_code,"city":ap.city or "Unknown Site","province":ap.province or "","country":ap.country or "","ap_model":ap.ap_model,"ip_address":ap.ip_address,"ethernet_mac":ap.ethernet_mac,"radio_mac":ap.radio_mac,"serial_number":ap.serial_number,"last_neighbor_detection":match["detected"] if match else None,"connected_switch":match["switch"] if match else None,"connected_interface":match["interface"] if match else None,"last_seen":ap.last_seen,"last_status_check":ap.last_status_check,"unknown_site":not bool(ap.city)}


def refresh_ap_status(db):
    """Persist each AP's online/offline status from the CDP-LLDP neighbors of all switches."""
    neighbors,collected_at=latest_neighbors(db);evidence_fresh=neighbor_evidence_is_fresh(collected_at);now=datetime.now(timezone.utc);aps=db.scalars(select(AccessPointInventory)).all();online=0
    for ap in aps:
        match=detected(ap,neighbors)
        ap.status="Online" if match else ("Offline" if evidence_fresh else "Unknown")
        ap.connected_switch=match["switch"] if match else None
        ap.connected_interface=match["interface"] if match else None
        ap.last_status_check=now
        if match:
            ap.last_seen=match["detected"] or now;online+=1
    db.commit()
    return {"total":len(aps),"online":online,"offline":sum(ap.status=="Offline" for ap in aps),"unknown":sum(ap.status=="Unknown" for ap in aps)}


async def ap_status_loop():
    """Check AP presence in switch CDP-LLDP neighbors every 2 hours (and once at startup)."""
    while True:
        try:
            with SessionLocal() as db: result=refresh_ap_status(db)
            logger.info("AP status refreshed: %s", result)
        except asyncio.CancelledError:
            raise
        except Exception as exc:
            logger.error("AP status refresh failed (%s)", type(exc).__name__)
        await asyncio.sleep(2*3600)


@router.get("")
def list_access_points(q:str="",site:str="",city:str="",province:str="",status:str="",model:str="",page:int=Query(1,ge=1),page_size:int=Query(50,ge=1,le=500),user=Depends(current_user),db:Session=Depends(session)):
    stmt=select(AccessPointInventory)
    if q:
        pattern=f"%{q}%";stmt=stmt.where(or_(AccessPointInventory.ap_name.ilike(pattern),AccessPointInventory.site_code.ilike(pattern),AccessPointInventory.city.ilike(pattern),AccessPointInventory.province.ilike(pattern),AccessPointInventory.serial_number.ilike(pattern),AccessPointInventory.ethernet_mac.ilike(pattern),AccessPointInventory.radio_mac.ilike(pattern),AccessPointInventory.ip_address.ilike(pattern)))
    for value,column in ((site,AccessPointInventory.site_code),(city,AccessPointInventory.city),(province,AccessPointInventory.province),(model,AccessPointInventory.ap_model)):
        if value: stmt=stmt.where(column==value)
    neighbors,collected_at=latest_neighbors(db);evidence_fresh=neighbor_evidence_is_fresh(collected_at);all_items=[row_json(x,neighbors,evidence_fresh) for x in db.scalars(stmt.order_by(AccessPointInventory.ap_name)).all()]
    if status: all_items=[x for x in all_items if x["status"]==status]
    start=(page-1)*page_size;serials=Counter(x["serial_number"] for x in all_items if x["serial_number"]);macs=Counter(m for x in all_items for m in (x["ethernet_mac"],x["radio_mac"]) if m)
    return {"items":all_items[start:start+page_size],"total":len(all_items),"page":page,"pages":max(1,(len(all_items)+page_size-1)//page_size),"summary":{"total":len(all_items),"online":sum(x["status"]=="Online" for x in all_items),"offline":sum(x["status"]=="Offline" for x in all_items),"unknown_site":sum(x["unknown_site"] for x in all_items),"duplicate_serials":sum(v>1 for v in serials.values()),"duplicate_macs":sum(v>1 for v in macs.values())}}


@router.post("/import/validate")
async def validate_import(file:UploadFile=File(...),user=Depends(administrator),db:Session=Depends(session)):
    started=time.monotonic();content=await file.read(MAX_AP_UPLOAD_BYTES + 1)
    if not file.filename or not file.filename.lower().endswith(".xlsx"): raise HTTPException(415,"Only .xlsx access-point exports are supported")
    _validate_workbook_bytes(content)
    try:
        ws=load_workbook(io.BytesIO(content),read_only=True,data_only=True).active
        # Some controller exports incorrectly declare <dimension ref="A1"> even
        # though the worksheet contains many columns. Streaming mode trusts that
        # metadata unless dimensions are reset and would otherwise read only A1.
        ws.reset_dimensions()
    except Exception: raise HTTPException(422,"The uploaded workbook is not a valid Excel file")
    headers=[str(x or "").strip() for x in next(ws.iter_rows(values_only=True))];missing=[x for x in REQUIRED if x not in headers]
    if missing: raise HTTPException(422,"Missing required columns: "+", ".join(missing))
    positions={name:headers.index(name) for name in REQUIRED};raw=[]
    for number,values in enumerate(ws.iter_rows(min_row=2,values_only=True),2):
        if len(raw) >= MAX_AP_IMPORT_ROWS:
            raise HTTPException(422, "The access-point workbook may contain at most 5,000 records")
        item={key:str(values[index] or "").strip() for key,index in positions.items()};item["row_number"]=number
        if any(item[x] for x in REQUIRED): raw.append(item)
    counts={field:Counter(x[field].lower() for x in raw if x[field]) for field in ("AP Name","Ethernet MAC","AP Radio MAC","Serial Number")};sites={x.site_code:x for x in db.scalars(select(Site)).all()};rows=[]
    for item in raw:
        errors=[];warnings=[];site_code=item["Site Tag"].upper();site=sites.get(site_code)
        if not item["AP Name"]: errors.append("AP Name is required")
        for field,label in (("Ethernet MAC","Ethernet MAC"),("AP Radio MAC","Radio MAC")):
            if item[field] and not MAC_RE.fullmatch(norm_mac(item[field])): errors.append(f"Invalid {label} format")
        if item["IP Address"]:
            try: ip_address(item["IP Address"])
            except ValueError: errors.append("Invalid IP address")
        for field,label in (("AP Name","AP name"),("Ethernet MAC","Ethernet MAC"),("AP Radio MAC","Radio MAC"),("Serial Number","serial number")):
            if item[field] and counts[field][item[field].lower()]>1: errors.append(f"Duplicate {label}")
        if not site: warnings.append("Unknown Site")
        status="Error" if errors else ("Warning" if warnings else "Ready")
        rows.append({"row_number":item["row_number"],"ap_name":item["AP Name"],"ap_model":item["AP Model"],"ip_address":item["IP Address"],"radio_mac":display_mac(item["AP Radio MAC"]),"ethernet_mac":display_mac(item["Ethernet MAC"]),"serial_number":item["Serial Number"],"site_code":site_code,"city":site.city if site else None,"province":site.province_region if site else None,"country":site.country if site else None,"validation_status":status,"validation_message":"; ".join(errors+warnings)})
    job=AccessPointImport(id=str(uuid.uuid4()),imported_by=user["sub"],file_name=file.filename,file_checksum=hashlib.sha256(content).hexdigest(),record_count=len(rows),validation_errors=sum(x["validation_status"]!="Ready" for x in rows),duration_ms=int((time.monotonic()-started)*1000),rows=rows);db.add(job);db.commit()
    return {"job_id":job.id,"rows":rows,"summary":{"total":len(rows),"ready":sum(x["validation_status"]=="Ready" for x in rows),"warnings":sum(x["validation_status"]=="Warning" for x in rows),"errors":sum(x["validation_status"]=="Error" for x in rows),"review":sum(x["validation_status"]!="Ready" for x in rows)}}


@router.post("/import/{job_id}/commit")
def commit_import(job_id:str,mode:str="replace",user=Depends(administrator),db:Session=Depends(session)):
    if mode not in ("replace","merge"): raise HTTPException(422,"Import mode must be replace or merge")
    job=db.get(AccessPointImport,job_id)
    if not job or job.status!="Validated": raise HTTPException(404,"Validated import not found")
    if any(x.get("validation_status")=="Error" for x in job.rows): raise HTTPException(422,"Resolve validation errors before importing access points")
    started=time.monotonic()
    if mode=="replace": db.execute(delete(AccessPointInventory))
    imported=0
    for value in job.rows:
        existing=db.scalar(select(AccessPointInventory).where(func.lower(AccessPointInventory.ap_name)==value["ap_name"].lower())) if mode=="merge" else None
        fields={k:value.get(k) for k in ("ap_name","site_code","city","province","country","ap_model","ip_address","ethernet_mac","radio_mac","serial_number")}
        if existing:
            for key,val in fields.items(): setattr(existing,key,val)
            existing.import_date=datetime.now(timezone.utc);existing.imported_by=user["sub"]
        else: db.add(AccessPointInventory(**fields,imported_by=user["sub"]));imported+=1
    job.import_mode=mode;job.status="Imported";job.duration_ms=(job.duration_ms or 0)+int((time.monotonic()-started)*1000)
    db.add(AuditEvent(actor=user["sub"],action="access_points.import",target=f"access_point_import:{job.id}",details={"file_name":job.file_name,"checksum":job.file_checksum,"records":job.record_count,"validation_errors":job.validation_errors,"mode":mode,"duration_ms":job.duration_ms}));db.commit()
    return {"status":"Imported","records":job.record_count,"created":imported,"mode":mode}


@router.get("/imports")
def import_history(user=Depends(current_user),db:Session=Depends(session)):
    return [{"id":x.id,"import_date":x.created_at,"imported_by":x.imported_by,"file_name":x.file_name,"record_count":x.record_count,"validation_errors":x.validation_errors,"duration_ms":x.duration_ms,"import_mode":x.import_mode,"status":x.status,"file_checksum":x.file_checksum} for x in db.scalars(select(AccessPointImport).order_by(AccessPointImport.created_at.desc()).limit(50)).all()]
