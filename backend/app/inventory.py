"""Device Fleet inventory router: the local, authoritative device registry.

Endpoints here let administrators add, import (CSV/Excel), edit, map, and remove the
switches/routers that make up Device Fleet — the scope every analytic in the app is
restricted to. Mapping a device links it to a LogicMonitor device id; refreshing a
device pulls a fresh read-only snapshot. Also serves site codes, zones, and device
types (the building blocks of the generated device names). Nothing here ever writes
to LogicMonitor — only to this application's own inventory tables.
"""
import csv
import io
import re
import uuid
from datetime import datetime, timezone
from ipaddress import ip_address
from pathlib import Path

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import FileResponse, JSONResponse
from openpyxl import load_workbook
from pydantic import BaseModel, Field, field_validator
from sqlalchemy import func, or_, select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from .auth import administrator, current_user
from .collection import collect_logicmonitor_device, latest
from .db import AuditEvent, Device, ImportJob, InventoryDevice, InventoryDeviceType, Site, Snapshot, Zone, session
from .logicmonitor import LogicMonitorClient
from .switch_refresh import refresh_switch_locked

router = APIRouter(prefix="/api/v1")
SITE_CODE = re.compile(r"^[A-Z0-9]{2,10}$")
ZONE_CODE = re.compile(r"^Z[0-9]{2}$")
TYPE_CODE = re.compile(r"^[A-Z0-9]{2,10}$")
BOOLS = {"true": True, "yes": True, "1": True, "enabled": True, "false": False, "no": False, "0": False, "disabled": False}
CRITICALITIES = {"Low", "Medium", "High", "Critical"}

DEFAULT_SITES = [
    ("CA05","Delta","BC"),("CA10","North Vancouver","BC"),("CAD01","Edmonton","AB"),("CA14","Calgary","AB"),("CAD14","Calgary","AB"),
    ("CA08","Winnipeg","MB"),("CAD08","Winnipeg","MB"),("CA15","Guelph","ON"),("MCA","Mississauga","ON"),("YRK","Toronto","ON"),
    ("CAD15","Mississauga","ON"),("CA06","Dartmouth","NS"),("CA07","Paradise","NL"),("CA20","Terrebonne","MO"),("CAD20","Montreal","QC"),
    ("CAD02","Ottawa","OT"),("CAD03","Quebec City","QC"),("CAD04","London","LO"),("CAD06","Halifax","NX")]
DEFAULT_TYPES = [("DSW","Distribution Switch","Distribution"),("ASW","Access Switch","Access"),("RTR","Router","Router")]


def seed_inventory(db: Session):
    if not db.scalar(select(func.count(Site.id))):
        db.add_all([Site(site_code=c, city=city, province_region=p, country="Canada", description=f"{city}, {p}") for c,city,p in DEFAULT_SITES])
    if not db.scalar(select(func.count(Zone.id))):
        db.add_all([Zone(zone_code=f"Z{i:02d}", description=f"Zone {i}", display_order=i) for i in range(1,10)])
    if not db.scalar(select(func.count(InventoryDeviceType.id))):
        db.add_all([InventoryDeviceType(type_code=c,description=d,derived_role=r,display_order=i) for i,(c,d,r) in enumerate(DEFAULT_TYPES,1)])
    else:
        for wireless in db.scalars(select(InventoryDeviceType).where(InventoryDeviceType.type_code.in_(("AP","WAP")))).all():
            if not db.scalar(select(func.count(InventoryDevice.id)).where(InventoryDevice.device_type_id==wireless.id)): db.delete(wireless)
            else: wireless.enabled=False
    db.commit()


def audit(db, actor, action, entity, entity_id=None, old=None, new=None, result="success"):
    db.add(AuditEvent(actor=actor, action=action, target=f"{entity}:{entity_id or ''}", details={"entity":entity,"entity_id":entity_id,"old":old,"new":new,"result":result}))


class SiteIn(BaseModel):
    site_code: str = Field(min_length=2,max_length=10)
    city: str = Field(min_length=1,max_length=120)
    province_region: str = Field(default="",max_length=80)
    country: str = Field(default="Canada",min_length=1,max_length=80)
    business_unit: str = Field(default="Unassigned",max_length=40)
    description: str | None = Field(default=None,max_length=500)
    enabled: bool = True
    @field_validator("site_code")
    @classmethod
    def valid_code(cls,v):
        v=v.strip().upper()
        if not SITE_CODE.fullmatch(v): raise ValueError("Use uppercase letters and numbers only")
        return v
    @field_validator("city","province_region","country")
    @classmethod
    def trim(cls,v): return v.strip()


class ZoneIn(BaseModel):
    zone_code: str
    description: str
    display_order: int = 0
    enabled: bool = True
    @field_validator("zone_code")
    @classmethod
    def valid_code(cls,v):
        v=v.strip().upper()
        if not ZONE_CODE.fullmatch(v): raise ValueError("Zone must use Z plus two digits")
        return v


class DeviceTypeIn(BaseModel):
    type_code: str
    description: str
    derived_role: str
    enabled: bool = True
    display_order: int = 0
    @field_validator("type_code")
    @classmethod
    def valid_code(cls,v):
        v=v.strip().upper()
        if not TYPE_CODE.fullmatch(v): raise ValueError("Invalid type code")
        return v


class NamePreview(BaseModel):
    site_code: str
    zone: str
    device_type: str
    device_number: str


class InventoryIn(NamePreview):
    criticality: str = "Medium"
    management_ip: str | None = None
    logicmonitor_device_id: int | None = None
    logicmonitor_display_name: str | None = None
    logicmonitor_match_status: str = "Pending LogicMonitor Match"
    logicmonitor_match_confidence: str | None = None
    model: str | None = None
    manufacturer: str | None = None
    os_version: str | None = None
    logicmonitor_groups: list = Field(default_factory=list)
    critical_interfaces: list = Field(default_factory=list)
    ignored_interfaces: list = Field(default_factory=list)
    tags: list = Field(default_factory=list)
    notes: str | None = None
    enabled: bool = True
    @field_validator("device_number")
    @classmethod
    def number(cls,v):
        v=str(v).strip().zfill(2)
        if not re.fullmatch(r"[0-9]{2}",v): raise ValueError("Device number must be two digits")
        return v
    @field_validator("criticality")
    @classmethod
    def criticality_valid(cls,v):
        if v not in CRITICALITIES: raise ValueError("Invalid criticality")
        return v
    @field_validator("management_ip")
    @classmethod
    def ip_valid(cls,v):
        if v: ip_address(v)
        return v


def refs(db, body):
    site=db.scalar(select(Site).where(Site.site_code==body.site_code.strip().upper()))
    zone=db.scalar(select(Zone).where(Zone.zone_code==body.zone.strip().upper()))
    dtype=db.scalar(select(InventoryDeviceType).where(InventoryDeviceType.type_code==body.device_type.strip().upper()))
    if not site or not site.enabled: raise HTTPException(422,"Site code does not exist or is disabled")
    if not zone or not zone.enabled: raise HTTPException(422,"Zone does not exist or is disabled")
    if not dtype or not dtype.enabled: raise HTTPException(422,"Device type does not exist or is disabled")
    return site,zone,dtype


def generated(site,zone,dtype,number): return f"{site.site_code}-{zone.zone_code}-{dtype.type_code}-{number}"


CHASSIS_PID = re.compile(r"^(?:C\d{4}[A-Z0-9-]*|WS-C\d{4}[A-Z0-9-]*)$", re.I)
STACK_MEMBER = re.compile(r"StackPort\s*(\d+)\s*/", re.I)


def physical_switch_count(details):
    """Count physical chassis from authoritative collected inventory evidence."""
    rows = ((details or {}).get("tables") or {}).get("Inventory") or []
    chassis_serials = set()
    stack_members = set()
    for row in rows:
        description = str(row.get("Description") or "").strip()
        pid = str(row.get("PID") or "").strip()
        serial = str(row.get("Serial Number") or "").strip().upper()
        model = pid if CHASSIS_PID.fullmatch(pid) else (description if CHASSIS_PID.fullmatch(description) else "")
        if model and serial and serial.lower() != "not available from logicmonitor":
            chassis_serials.add(serial)
        match = STACK_MEMBER.search(description)
        if match:
            stack_members.add(int(match.group(1)))
    if chassis_serials:
        return len(chassis_serials), "Inventory chassis serials"
    if stack_members:
        return len(stack_members), "Stack-port member evidence"
    return None, "Not monitored"


def switch_evidence(db, d):
    if d.device_type.type_code not in {"DSW", "ASW"}:
        return None, "Not applicable"
    legacy = db.scalar(select(Device).where(Device.lm_device_id == d.logicmonitor_device_id)) if d.logicmonitor_device_id else None
    snapshot = db.scalar(select(Snapshot).where(Snapshot.device_id == legacy.id).order_by(Snapshot.collected_at.desc()).limit(1)) if legacy else None
    return physical_switch_count(snapshot.details if snapshot else None)


def device_json(d, db=None):
    count, source = switch_evidence(db, d) if db else (None, "Not loaded")
    return {"id":d.id,"generated_name":d.generated_name,"site_code":d.site.site_code,"city":d.site.city,"province_region":d.site.province_region,"zone":d.zone.zone_code,"device_type":d.device_type.type_code,"device_type_description":d.device_type.description,"device_number":d.device_number,"physical_switch_count":count,"physical_switch_count_source":source,"role":d.role,"model":d.model,"management_ip":d.management_ip,"logicmonitor_device_id":d.logicmonitor_device_id,"logicmonitor_display_name":d.logicmonitor_display_name,"logicmonitor_match_status":d.logicmonitor_match_status,"logicmonitor_match_confidence":d.logicmonitor_match_confidence,"criticality":d.criticality,"enabled":d.enabled,"last_synchronized":d.last_logicmonitor_sync,"manufacturer":d.manufacturer,"os_version":d.os_version,"logicmonitor_groups":d.logicmonitor_groups,"critical_interfaces":d.critical_interfaces,"ignored_interfaces":d.ignored_interfaces,"tags":d.tags,"notes":d.notes}


def readable_uptime(seconds):
    if not isinstance(seconds,(int,float)): return None
    total=max(0,int(seconds));days,remainder=divmod(total,86400);hours,remainder=divmod(remainder,3600);minutes,_=divmod(remainder,60)
    years,days=divmod(days,365);weeks,days=divmod(days,7)
    parts=[]
    for value,label in ((years,"year"),(weeks,"week"),(days,"day"),(hours,"hour"),(minutes,"minute")):
        if value: parts.append(f"{value} {label}{'' if value==1 else 's'}")
    return ", ".join(parts) or "0 minutes"


async def live_uptime(row,client):
    if not row.logicmonitor_device_id: return {"raw_seconds":None,"display":"Pending LogicMonitor match","last_polled":None}
    applied=await client.applied_datasources(row.logicmonitor_device_id)
    ds=next((x for x in applied if (x.get("dataSourceName") or x.get("name"))=="SNMP_Host_Uptime"),None)
    if not ds: return {"raw_seconds":None,"display":"Not monitored","last_polled":None}
    instances=await client.instances(row.logicmonitor_device_id,int(ds["id"]))
    if not instances: return {"raw_seconds":None,"display":"No uptime instance","last_polled":None}
    end=int(datetime.now(timezone.utc).timestamp());data=await client.instance_data(row.logicmonitor_device_id,int(ds["id"]),int(instances[0]["id"]),end-3600,end);value=latest(data)
    raw=next((value.get(k) for k in ("Uptime","uptime","sysUpTime") if isinstance(value.get(k),(int,float))),None)
    return {"raw_seconds":raw,"display":readable_uptime(raw),"last_polled":datetime.now(timezone.utc).isoformat() if raw is not None else None}


@router.get("/settings/sites")
def sites(q:str="", user=Depends(current_user), db:Session=Depends(session)):
    rows=db.scalars(select(Site).where(or_(Site.site_code.ilike(f"%{q}%"),Site.city.ilike(f"%{q}%"))).order_by(Site.site_code)).all()
    response=[]
    for x in rows:
        switches=db.scalars(select(InventoryDevice).join(InventoryDeviceType).where(InventoryDevice.site_id==x.id,InventoryDeviceType.type_code.in_(("DSW","ASW")))).all()
        evidence=[switch_evidence(db,d)[0] for d in switches]
        response.append({"id":x.id,"site_code":x.site_code,"city":x.city,"province_region":x.province_region,"country":x.country,"business_unit":x.business_unit,"description":x.description,"enabled":x.enabled,"number_of_devices":sum(value for value in evidence if value is not None),"switch_count_pending":sum(value is None for value in evidence),"last_modified":x.updated_at,"modified_by":x.updated_by})
    return response


@router.post("/settings/sites",status_code=201)
def add_site(body:SiteIn,user=Depends(administrator),db:Session=Depends(session)):
    if db.scalar(select(Site).where(or_(Site.site_code==body.site_code,func.lower(Site.city)==body.city.lower()))): raise HTTPException(409,"Duplicate site-code or city mapping")
    row=Site(**body.model_dump(),created_by=user["sub"],updated_by=user["sub"]);db.add(row);db.flush();audit(db,user["sub"],"site.create","site",row.id,new=body.model_dump());db.commit();db.refresh(row);return {"id":row.id,**body.model_dump()}


@router.put("/settings/sites/{item_id}")
def edit_site(item_id:int,body:SiteIn,user=Depends(administrator),db:Session=Depends(session)):
    row=db.get(Site,item_id)
    if not row: raise HTTPException(404,"Site code not found")
    duplicate=db.scalar(select(Site).where(Site.id!=item_id,or_(Site.site_code==body.site_code,func.lower(Site.city)==body.city.lower())))
    if duplicate: raise HTTPException(409,"Duplicate site-code or city mapping")
    old={k:getattr(row,k) for k in body.model_fields};[setattr(row,k,v) for k,v in body.model_dump().items()];row.updated_by=user["sub"];audit(db,user["sub"],"site.update","site",row.id,old,body.model_dump());db.commit();return {"id":row.id,**body.model_dump()}


@router.delete("/settings/sites/{item_id}",status_code=204)
def delete_site(item_id:int,user=Depends(administrator),db:Session=Depends(session)):
    row=db.get(Site,item_id)
    if not row: raise HTTPException(404,"Site code not found")
    if db.scalar(select(func.count(InventoryDevice.id)).where(InventoryDevice.site_id==item_id)): raise HTTPException(409,"Site is assigned to devices; disable it or reassign devices first")
    audit(db,user["sub"],"site.delete","site",item_id,old={"site_code":row.site_code});db.delete(row);db.commit()


@router.get("/settings/zones")
def zones(user=Depends(current_user),db:Session=Depends(session)): return [{"id":x.id,"zone_code":x.zone_code,"description":x.description,"display_order":x.display_order,"enabled":x.enabled} for x in db.scalars(select(Zone).order_by(Zone.display_order)).all()]
@router.post("/settings/zones",status_code=201)
def add_zone(body:ZoneIn,user=Depends(administrator),db:Session=Depends(session)): row=Zone(**body.model_dump());db.add(row);db.flush();audit(db,user["sub"],"zone.create","zone",row.id,new=body.model_dump());db.commit();return {"id":row.id,**body.model_dump()}
@router.put("/settings/zones/{item_id}")
def edit_zone(item_id:int,body:ZoneIn,user=Depends(administrator),db:Session=Depends(session)): row=db.get(Zone,item_id) or (_ for _ in ()).throw(HTTPException(404,"Zone not found"));[setattr(row,k,v) for k,v in body.model_dump().items()];audit(db,user["sub"],"zone.update","zone",item_id,new=body.model_dump());db.commit();return {"id":row.id,**body.model_dump()}
@router.delete("/settings/zones/{item_id}",status_code=204)
def delete_zone(item_id:int,user=Depends(administrator),db:Session=Depends(session)):
    row=db.get(Zone,item_id)
    if not row: raise HTTPException(404,"Zone not found")
    if db.scalar(select(func.count(InventoryDevice.id)).where(InventoryDevice.zone_id==item_id)): raise HTTPException(409,"Zone is assigned to devices; disable it or reassign devices first")
    audit(db,user["sub"],"zone.delete","zone",item_id,old={"zone_code":row.zone_code});db.delete(row);db.commit()


@router.get("/settings/device-types")
def device_types(user=Depends(current_user),db:Session=Depends(session)): return [{"id":x.id,"type_code":x.type_code,"description":x.description,"derived_role":x.derived_role,"enabled":x.enabled,"display_order":x.display_order} for x in db.scalars(select(InventoryDeviceType).where(InventoryDeviceType.type_code.not_in(("AP","WAP"))).order_by(InventoryDeviceType.display_order)).all()]
@router.post("/settings/device-types",status_code=201)
def add_type(body:DeviceTypeIn,user=Depends(administrator),db:Session=Depends(session)):
    if body.type_code in ("AP","WAP"): raise HTTPException(422,"Wireless access points must be imported through Access Point Inventory")
    row=InventoryDeviceType(**body.model_dump());db.add(row);db.commit();db.refresh(row);return {"id":row.id,**body.model_dump()}
@router.put("/settings/device-types/{item_id}")
def edit_type(item_id:int,body:DeviceTypeIn,user=Depends(administrator),db:Session=Depends(session)):
    if body.type_code in ("AP","WAP"): raise HTTPException(422,"Wireless access points must be imported through Access Point Inventory")
    row=db.get(InventoryDeviceType,item_id) or (_ for _ in ()).throw(HTTPException(404,"Device type not found"));[setattr(row,k,v) for k,v in body.model_dump().items()];audit(db,user["sub"],"device_type.update","device_type",item_id,new=body.model_dump());db.commit();return {"id":row.id,**body.model_dump()}
@router.delete("/settings/device-types/{item_id}",status_code=204)
def delete_type(item_id:int,user=Depends(administrator),db:Session=Depends(session)):
    row=db.get(InventoryDeviceType,item_id)
    if not row: raise HTTPException(404,"Device type not found")
    if row.type_code in ("AP","WAP"): raise HTTPException(422,"Wireless access points are managed in Access Point Inventory")
    if db.scalar(select(func.count(InventoryDevice.id)).where(InventoryDevice.device_type_id==item_id)): raise HTTPException(409,"Device type is assigned to devices; disable it or reassign devices first")
    audit(db,user["sub"],"device_type.delete","device_type",item_id,old={"type_code":row.type_code});db.delete(row);db.commit()


@router.post("/inventory/preview-name")
def preview_name(body:NamePreview,user=Depends(current_user),db:Session=Depends(session)):
    site,zone,dtype=refs(db,body);number=str(body.device_number).zfill(2)
    if not re.fullmatch(r"[0-9]{2}",number): raise HTTPException(422,"Device number must be two digits")
    name=generated(site,zone,dtype,number)
    return {"generated_name":name,"city":site.city,"role":dtype.derived_role,"unique":not bool(db.scalar(select(InventoryDevice.id).where(InventoryDevice.generated_name==name)))}


@router.get("/inventory/devices")
def inventory_devices(q:str="",site:str|None=None,city:str|None=None,zone:str|None=None,device_type:str|None=None,role:str|None=None,model:str|None=None,match_status:str|None=None,criticality:str|None=None,enabled:bool|None=None,sort:str="generated_name",direction:str="asc",page:int=Query(1,ge=1),page_size:int=Query(25,ge=1,le=200),user=Depends(current_user),db:Session=Depends(session)):
    stmt=select(InventoryDevice).join(Site).join(Zone).join(InventoryDeviceType)
    if q: stmt=stmt.where(or_(InventoryDevice.generated_name.ilike(f"%{q}%"),InventoryDevice.management_ip.ilike(f"%{q}%")))
    for value,column in [(site,Site.site_code),(city,Site.city),(zone,Zone.zone_code),(device_type,InventoryDeviceType.type_code),(role,InventoryDevice.role),(model,InventoryDevice.model),(match_status,InventoryDevice.logicmonitor_match_status),(criticality,InventoryDevice.criticality),(enabled,InventoryDevice.enabled)]:
        if value not in (None, ""): stmt=stmt.where(column==value)
    sort_map={"generated_name":InventoryDevice.generated_name,"site":Site.site_code,"city":Site.city,"zone":Zone.zone_code,"device_type":InventoryDeviceType.type_code,"model":InventoryDevice.model,"last_synchronized":InventoryDevice.last_logicmonitor_sync}
    stmt=stmt.order_by((sort_map.get(sort,InventoryDevice.generated_name).desc() if direction=="desc" else sort_map.get(sort,InventoryDevice.generated_name).asc()))
    all_rows=db.scalars(stmt).all();start=(page-1)*page_size
    total_physical=sum((switch_evidence(db,d)[0] or 1) for d in all_rows)
    return {"items":[device_json(x,db) for x in all_rows[start:start+page_size]],"total":len(all_rows),"total_physical":total_physical,"page":page,"page_size":page_size,"pages":max(1,(len(all_rows)+page_size-1)//page_size)}


@router.get("/inventory/devices/{item_id}")
def inventory_device(item_id:int,user=Depends(current_user),db:Session=Depends(session)): row=db.get(InventoryDevice,item_id);return device_json(row,db) if row else (_ for _ in ()).throw(HTTPException(404,"Device not found"))


@router.get("/inventory/devices/{item_id}/uptime")
async def inventory_uptime(item_id:int,user=Depends(current_user),db:Session=Depends(session)):
    row=db.get(InventoryDevice,item_id)
    if not row: raise HTTPException(404,"Device not found")
    try: return await live_uptime(row,LogicMonitorClient())
    except Exception as exc: return JSONResponse(status_code=502,content={"detail":"LogicMonitor uptime lookup failed","category":type(exc).__name__})


@router.post("/inventory/devices/{item_id}/open-detail")
async def open_inventory_detail(item_id:int,user=Depends(current_user),db:Session=Depends(session)):
    row=db.get(InventoryDevice,item_id)
    if not row: raise HTTPException(404,"Device not found")
    if not row.logicmonitor_device_id:
        raise HTTPException(409,"Device is not mapped to LogicMonitor")
    try:
        legacy=await refresh_switch_locked(db,row,user["sub"])
    except Exception as exc:
        db.rollback()
        raise HTTPException(502,detail={"message":"LogicMonitor switch refresh failed","category":type(exc).__name__})
    audit(db,user["sub"],"inventory.open_detail","inventory_device",row.id,new={"legacy_device_id":legacy.id});db.commit();db.refresh(legacy)
    return {"device":{"id":legacy.id,"hostname":legacy.hostname,"management_ip":legacy.management_ip,"site":legacy.site,"role":legacy.role,"criticality":legacy.criticality,"device_type":legacy.device_type,"model":legacy.model,"active":legacy.active,"notes":legacy.notes,"lm_device_id":legacy.lm_device_id,"match_status":legacy.match_status}}


def save_device(body:InventoryIn,user,db,row=None,commit=True):
    site,zone,dtype=refs(db,body);name=generated(site,zone,dtype,body.device_number)
    conflict=db.scalar(select(InventoryDevice).where(InventoryDevice.id!=(row.id if row else 0),or_(InventoryDevice.generated_name==name,*([InventoryDevice.management_ip==body.management_ip] if body.management_ip else []),*([InventoryDevice.logicmonitor_device_id==body.logicmonitor_device_id] if body.logicmonitor_device_id else []))))
    if conflict: raise HTTPException(409,f"{name} conflicts with existing record {conflict.generated_name}")
    values=body.model_dump(exclude={"site_code","zone","device_type"})|{"site_id":site.id,"zone_id":zone.id,"device_type_id":dtype.id,"generated_name":name,"role":dtype.derived_role,"updated_by":user["sub"]}
    if row:
        old=device_json(row);[setattr(row,k,v) for k,v in values.items()];action="device.update"
    else:
        row=InventoryDevice(**values,created_by=user["sub"]);db.add(row);old=None;action="device.create"
    try: db.flush()
    except IntegrityError: db.rollback();raise HTTPException(409,"Generated name, management IP, or LogicMonitor ID already exists")
    audit(db,user["sub"],action,"inventory_device",row.id,old,values)
    if commit: db.commit();db.refresh(row)
    return device_json(row)


@router.post("/inventory/devices",status_code=201)
def create_inventory_device(body:InventoryIn,user=Depends(administrator),db:Session=Depends(session)): return save_device(body,user,db)
@router.put("/inventory/devices/{item_id}")
def update_inventory_device(item_id:int,body:InventoryIn,user=Depends(administrator),db:Session=Depends(session)): row=db.get(InventoryDevice,item_id) or (_ for _ in ()).throw(HTTPException(404,"Device not found"));return save_device(body,user,db,row)
@router.delete("/inventory/devices/{item_id}",status_code=204)
def delete_inventory_device(item_id:int,user=Depends(administrator),db:Session=Depends(session)): row=db.get(InventoryDevice,item_id) or (_ for _ in ()).throw(HTTPException(404,"Device not found"));audit(db,user["sub"],"device.delete","inventory_device",item_id,old=device_json(row));db.delete(row);db.commit()


@router.post("/inventory/logicmonitor-lookup")
async def lm_lookup(body:NamePreview,user=Depends(current_user),db:Session=Depends(session)):
    site,zone,dtype=refs(db,body);name=generated(site,zone,dtype,str(body.device_number).zfill(2));client=LogicMonitorClient()
    try:
        candidates=[]
        for field in ("displayName","name"):
            payload=await client.get("/santaba/rest/device/devices",{"size":20,"filter":f'{field}:"{name}"'})
            for item in client.body(payload).get("items",[]):
                if int(item["id"]) not in {int(x["id"]) for x in candidates}: candidates.append(item)
        exact=[x for x in candidates if name.lower() in {str(x.get("displayName","")).lower(),str(x.get("name","")).lower(),str(x.get("name","")).lower().split(".")[0]}]
        if not exact: return {"status":"Not Found in LogicMonitor","generated_name":name,"candidates":[]}
        results=[]
        for item in exact:
            props=await client.properties(int(item["id"]));model_text=props.get("auto.endpoint.model","");model_match=re.search(r"\bC\d{4}(?:-[A-Z0-9]+)+\b",model_text,re.I)
            legacy=db.scalar(select(Device).where(Device.lm_device_id==int(item["id"])))
            raw_model=(legacy.model if legacy and legacy.model else None) or (model_match.group(0).upper() if model_match else (props.get("system.model") or props.get("auto.endpoint.model")))
            model=(str(raw_model)[:100] if raw_model else None)
            raw_groups=item.get("hostGroupIds",[])
            groups=[x.strip() for x in raw_groups.split(",") if x.strip()] if isinstance(raw_groups,str) else (raw_groups or [])
            results.append({"logicmonitor_device_id":item.get("id"),"display_name":item.get("displayName"),"hostname":item.get("name"),"management_ip":item.get("name") if re.fullmatch(r"[0-9a-fA-F:.]+",str(item.get("name",""))) else props.get("system.ips"),"model":model,"manufacturer":props.get("system.manufacturer") or "Cisco","os_version":props.get("system.version"),"groups":groups,"status":item.get("hostStatus"),"last_data_timestamp":item.get("lastDataTime")})
        return {"status":"Matched" if len(results)==1 else "Ambiguous","generated_name":name,"match_confidence":"Exact" if len(results)==1 else None,"candidates":results}
    except Exception as exc:
        return JSONResponse(status_code=502,content={"detail":"LogicMonitor lookup failed","category":type(exc).__name__})


@router.post("/inventory/{item_id}/refresh")
async def refresh_inventory(item_id:int,user=Depends(administrator),db:Session=Depends(session)):
    row=db.get(InventoryDevice,item_id) or (_ for _ in ()).throw(HTTPException(404,"Device not found"));body=NamePreview(site_code=row.site.site_code,zone=row.zone.zone_code,device_type=row.device_type.type_code,device_number=row.device_number);result=await lm_lookup(body,user,db)
    if isinstance(result,dict) and result.get("status")=="Matched":
        match=result["candidates"][0];row.management_ip=match.get("management_ip") or row.management_ip;row.model=match.get("model") or row.model;row.os_version=match.get("os_version") or row.os_version;row.logicmonitor_device_id=match["logicmonitor_device_id"];row.logicmonitor_display_name=match["display_name"];row.logicmonitor_groups=match.get("groups",[]);row.logicmonitor_match_status="Matched";row.last_logicmonitor_sync=datetime.now(timezone.utc);audit(db,user["sub"],"logicmonitor.refresh","inventory_device",row.id,new={"status":"Matched"});db.commit()
    return result


def parse_upload(file):
    raw=file.file.read()
    if len(raw)>10*1024*1024: raise HTTPException(413,"Maximum upload size is 10 MB")
    suffix=Path(file.filename or "").suffix.lower()
    if suffix==".csv": return list(csv.DictReader(io.StringIO(raw.decode("utf-8-sig"))))
    if suffix==".xlsx":
        wb=load_workbook(io.BytesIO(raw),read_only=True,data_only=True);ws=wb["Device Inventory"] if "Device Inventory" in wb.sheetnames else wb.active;values=list(ws.values);headers=[str(x or "") for x in values[0]];return [dict(zip(headers,row)) for row in values[1:] if any(x is not None for x in row)]
    raise HTTPException(415,"Only .xlsx and .csv files are supported")


def validate_rows(rows,db):
    if len(rows)>5000: raise HTTPException(422,"Maximum 5,000 rows")
    output=[];seen=set()
    for number,raw in enumerate(rows,2):
        safe={str(k):("'"+v if isinstance(v,str) and v[:1] in "=+-@" else v) for k,v in raw.items()};errors=[]
        try:
            enabled=BOOLS.get(str(safe.get("Enabled","true")).strip().lower())
            if enabled is None: errors.append("Enabled must be true/false")
            body=InventoryIn(site_code=str(safe.get("SiteCode","")).strip().upper(),zone=str(safe.get("Zone","")).strip().upper(),device_type=str(safe.get("DeviceType","")).strip().upper(),device_number=str(safe.get("DeviceNumber","")).strip().zfill(2),criticality=str(safe.get("Criticality","Medium")).strip(),critical_interfaces=[x.strip() for x in str(safe.get("CriticalInterfaces","")).split(",") if x.strip()],ignored_interfaces=[x.strip() for x in str(safe.get("IgnoreInterfaces","")).split(",") if x.strip()],tags=[x.strip() for x in str(safe.get("Tags","")).split(",") if x.strip()],notes=str(safe.get("Notes","") or ""),enabled=enabled if enabled is not None else True)
            site,zone,dtype=refs(db,body);name=generated(site,zone,dtype,body.device_number)
            if name in seen: errors.append("Duplicate generated name in upload")
            if db.scalar(select(InventoryDevice.id).where(InventoryDevice.generated_name==name)): errors.append("Device already exists")
            seen.add(name)
            item=body.model_dump()|{"row_number":number,"generated_name":name,"city":site.city,"role":dtype.derived_role}
        except Exception as exc:
            item={"row_number":number,"generated_name":"","city":"","role":"",**safe};errors.append(getattr(exc,"detail",str(exc)))
        item["validation_status"]="Ready" if not errors else ("Duplicate" if any("uplicate" in str(x) or "exists" in str(x) for x in errors) else "Error");item["validation_message"]="; ".join(map(str,errors));output.append(item)
    return output


@router.get("/inventory/import/template")
def import_template(user=Depends(current_user)):
    path=Path(__file__).parent/"assets"/"Device_Inventory_Import_Template.xlsx"
    if not path.is_file(): raise HTTPException(503,"Import template is not installed")
    return FileResponse(path,filename=path.name,media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@router.post("/inventory/import/validate")
def validate_import(file:UploadFile=File(...),mode:str="valid_rows_only",user=Depends(administrator),db:Session=Depends(session)):
    rows=validate_rows(parse_upload(file),db);job=ImportJob(id=str(uuid.uuid4()),actor=user["sub"],mode=mode,rows=rows,summary={"total":len(rows),"ready":sum(x["validation_status"]=="Ready" for x in rows),"errors":sum(x["validation_status"]!="Ready" for x in rows)});db.add(job);audit(db,user["sub"],"bulk_import.validate","import_job",job.id,new=job.summary);db.commit();return {"job_id":job.id,"rows":rows,"summary":job.summary}


@router.post("/inventory/import/commit")
def commit_import(job_id:str,mode:str="valid_rows_only",user=Depends(administrator),db:Session=Depends(session)):
    job=db.get(ImportJob,job_id) or (_ for _ in ()).throw(HTTPException(404,"Import job not found"));invalid=[x for x in job.rows if x["validation_status"]!="Ready"]
    if mode=="all_or_nothing" and invalid: raise HTTPException(422,"All-or-nothing import rejected because rows contain errors")
    imported=0
    try:
        for x in job.rows:
            if x["validation_status"]!="Ready": continue
            body=InventoryIn(**{k:x[k] for k in InventoryIn.model_fields if k in x});save_device(body,user,db,commit=False);imported+=1
        job.status="Committed";job.summary=job.summary|{"imported":imported,"skipped":len(job.rows)-imported};audit(db,user["sub"],"bulk_import.commit","import_job",job.id,new=job.summary);db.commit()
    except Exception:
        db.rollback();audit(db,user["sub"],"bulk_import.failed","import_job",job.id,result="failed");db.commit();raise
    return job.summary


@router.get("/inventory/import/{job_id}")
def import_job(job_id:str,user=Depends(current_user),db:Session=Depends(session)): job=db.get(ImportJob,job_id) or (_ for _ in ()).throw(HTTPException(404,"Import job not found"));return {"job_id":job.id,"status":job.status,"summary":job.summary,"rows":job.rows}
@router.get("/inventory/import/{job_id}/errors")
def import_errors(job_id:str,user=Depends(current_user),db:Session=Depends(session)): job=db.get(ImportJob,job_id) or (_ for _ in ()).throw(HTTPException(404,"Import job not found"));rows=[x for x in job.rows if x["validation_status"]!="Ready"];text=io.StringIO();w=csv.DictWriter(text,fieldnames=sorted({k for x in rows for k in x}));w.writeheader();w.writerows(rows);return JSONResponse({"filename":f"{job_id}-errors.csv","content":text.getvalue()})
